# -*- coding: utf-8 -*-
"""
openpyxl 常用读写操作演示
运行: .venv/Scripts/python.exe demo_xlsx.py
生成: output/demo_xlsx.xlsx   (同时会读回并打印内容)
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

OUT_DIR = "output"
os.makedirs(OUT_DIR, exist_ok=True)
XLSX_PATH = os.path.join(OUT_DIR, "demo_xlsx.xlsx")


# ---------------------------------------------------------------
# 一、写入操作
# ---------------------------------------------------------------
def write_xlsx(path):
    wb = Workbook()

    # ===== 工作表 1: 数据 + 样式 =====
    ws = wb.active
    ws.title = "员工工资表"

    # 1. 写入表头
    headers = ["姓名", "部门", "工资", "入职年份"]
    ws.append(headers)   # append 按行追加

    # 2. 写入数据
    data = [
        ["张三", "研发部", 15000, 2020],
        ["李四", "市场部", 12000, 2021],
        ["王五", "人事部", 10000, 2019],
        ["赵六", "研发部", 18000, 2018],
        ["钱七", "市场部", 11000, 2022],
    ]
    for row in data:
        ws.append(row)

    # 3. 单元格样式: 字体 / 背景色 / 边框 / 对齐
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")   # 蓝色背景
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:   # 第一行(表头)应用样式
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if cell.column == 3:                        # C列(工资)右对齐
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"            # 千分位格式

    # 4. 合并单元格 + 单元格直接赋值
    ws["A8"] = "合计"
    ws["C8"] = "=SUM(C2:C6)"                            # 公式
    ws.merge_cells("A8:B8")
    ws["A8"].font = Font(bold=True)

    # 5. 列宽 / 行高 / 冻结窗格
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"                              # 冻结首行

    # ===== 工作表 2: 图表 =====
    ws2 = wb.create_sheet("工资图表")
    ws2.append(["部门", "平均工资"])
    ws2.append(["研发部", 16500])
    ws2.append(["市场部", 11500])
    ws2.append(["人事部", 10000])

    chart = BarChart()
    chart.title = "各部门平均工资"
    chart.type = "col"
    data_ref = Reference(ws2, min_col=2, min_row=1, max_row=4)  # 数据区(含表头)
    cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=4)  # 分类(部门名)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws2.add_chart(chart, "E2")                          # 图表放在 E2 起的位置

    wb.save(path)
    print(f"[写入完成] {path}  (工作表: {wb.sheetnames})")


# ---------------------------------------------------------------
# 二、读取操作
# ---------------------------------------------------------------
def read_xlsx(path):
    wb = load_workbook(path, data_only=True)   # data_only=True 取公式计算结果(需文件曾被 Excel 打开过)
    print(f"\n--- 工作表列表: {wb.sheetnames} ---")

    # 1. 按名称获取工作表
    ws = wb["员工工资表"]
    print(f"\n--- 读取「{ws.title}」 (最大行 {ws.max_row}, 最大列 {ws.max_column}) ---")

    # 2. 方式一: 遍历所有单元格
    print("\n[方式一] 遍历全部单元格:")
    for row in ws.iter_rows(min_row=1, max_row=7):
        values = [str(c.value) if c.value is not None else "" for c in row]
        print("  " + " | ".join(values))

    # 3. 方式二: values_only 直接拿值
    print("\n[方式二] values_only 读取: 只取前 3 行")
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        print("  " + str(row))

    # 4. 方式三: 定位读取指定单元格 (行列从 1 开始)
    print("\n[方式三] 指定单元格:")
    print(f"  B2 = {ws['B2'].value}")
    print(f"  C3 = {ws.cell(row=3, column=3).value}")

    # 5. 读取图表
    ws2 = wb["工资图表"]
    print(f"\n[图表] 「{ws2.title}」 中有 {len(ws2._charts)} 个图表")


if __name__ == "__main__":
    write_xlsx(XLSX_PATH)
    read_xlsx(XLSX_PATH)
